import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Carregar dados processados
df_filtered = pd.read_excel('/home/ubuntu/processed_data.xlsx', sheet_name='Dados Filtrados')
df_bairros = pd.read_excel('/home/ubuntu/processed_data.xlsx', sheet_name='Resumo Bairros')
df_sangue = pd.read_excel('/home/ubuntu/processed_data.xlsx', sheet_name='Resumo Tipos Sanguineos')

wb = Workbook()

# --- Configurações de Tema ---
THEME = {'primary': '1F4E79', 'light': 'D6E3F0', 'accent': '1F4E79'}
SERIF_FONT = 'Calibri'

def apply_header_style(cell):
    cell.font = Font(name=SERIF_FONT, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_border(cell):
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

# --- Planilha 1: Visão Geral (Overview) ---
ws_ov = wb.active
ws_ov.title = "Visão Geral"
ws_ov.sheet_view.showGridLines = False
ws_ov.column_dimensions['A'].width = 3

ws_ov['B2'] = "RELATÓRIO DE DOADORES - GOIÂNIA (2023-2025)"
ws_ov['B2'].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME['primary'])

ws_ov['B4'] = "Este relatório apresenta a análise de doadores residentes em Goiânia, filtrados a partir da base original."
ws_ov['B5'] = f"Data de Geração: {datetime.now().strftime('%d/%m/%Y')}"

# KPI Banner (Delight feature)
ws_ov['B7'] = "TOTAL DE DOADORES"
ws_ov['B7'].font = Font(bold=True)
ws_ov['B8'] = len(df_filtered)
ws_ov['B8'].font = Font(size=14, bold=True)
ws_ov['B8'].number_format = '#,##0'

ws_ov['D7'] = "BAIRROS ATENDIDOS"
ws_ov['D7'].font = Font(bold=True)
ws_ov['D8'] = df_filtered['BAIRRO'].nunique()
ws_ov['D8'].font = Font(size=14, bold=True)

# Links de Navegação
ws_ov['B11'] = "CONTEÚDO"
ws_ov['B11'].font = Font(size=12, bold=True)
ws_ov['B12'] = "1. Resumo por Bairro"
ws_ov['B12'].hyperlink = "#'Resumo por Bairro'!A1"
ws_ov['B13'] = "2. Tipos Sanguíneos"
ws_ov['B13'].hyperlink = "#'Tipos Sanguíneos'!A1"
ws_ov['B14'] = "3. Dados Completos"
ws_ov['B14'].hyperlink = "#'Dados Completos'!A1"

# --- Planilha 2: Resumo por Bairro ---
ws_b = wb.create_sheet("Resumo por Bairro")
ws_b.sheet_view.showGridLines = False
ws_b.column_dimensions['A'].width = 3
ws_b['B2'] = "Principais Bairros com Doadores"
ws_b['B2'].font = Font(size=14, bold=True)

# Tabela de Bairros
for r_idx, row in enumerate(dataframe_to_rows(df_bairros, index=False, header=True), 4):
    for c_idx, value in enumerate(row, 2):
        cell = ws_b.cell(row=r_idx, column=c_idx, value=value)
        apply_border(cell)
        if r_idx == 4:
            apply_header_style(cell)
        else:
            cell.alignment = Alignment(horizontal="left")
            if c_idx == 3:
                cell.number_format = '#,##0'

# Gráfico de Bairros
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Top 15 Bairros com Doadores"
chart1.y_axis.title = 'Quantidade'
chart1.x_axis.title = 'Bairro'
data = Reference(ws_b, min_col=3, min_row=4, max_row=4+len(df_bairros))
cats = Reference(ws_b, min_col=2, min_row=5, max_row=4+len(df_bairros))
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.legend = None
ws_b.add_chart(chart1, "E4")

# --- Planilha 3: Tipos Sanguíneos ---
ws_s = wb.create_sheet("Tipos Sanguíneos")
ws_s.sheet_view.showGridLines = False
ws_s.column_dimensions['A'].width = 3
ws_s['B2'] = "Distribuição por Tipo Sanguíneo"
ws_s['B2'].font = Font(size=14, bold=True)

# Tabela de Tipos Sanguíneos
for r_idx, row in enumerate(dataframe_to_rows(df_sangue, index=False, header=True), 4):
    for c_idx, value in enumerate(row, 2):
        cell = ws_s.cell(row=r_idx, column=c_idx, value=value)
        apply_border(cell)
        if r_idx == 4:
            apply_header_style(cell)
        else:
            cell.alignment = Alignment(horizontal="center")
            if c_idx == 3:
                cell.number_format = '#,##0'

# Gráfico de Pizza
chart2 = PieChart()
chart2.title = "Distribuição de Tipos Sanguíneos"
data = Reference(ws_s, min_col=3, min_row=4, max_row=4+len(df_sangue))
cats = Reference(ws_s, min_col=2, min_row=5, max_row=4+len(df_sangue))
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
ws_s.add_chart(chart2, "E4")

# --- Planilha 4: Dados Completos ---
ws_d = wb.create_sheet("Dados Completos")
ws_d.freeze_panes = "A2"
for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_d.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            apply_header_style(cell)
        if c_idx == 2: # Data
            cell.number_format = 'dd/mm/yyyy'

# Ajustar larguras de coluna
for ws in wb.worksheets:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50

wb.save('/home/ubuntu/doadores_goiania_final.xlsx')
print("Planilha final gerada: /home/ubuntu/doadores_goiania_final.xlsx")
